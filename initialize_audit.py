# routes/initialize_audit.py

import os
import json
import shutil
from fastapi import APIRouter, Form
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from datetime import datetime

router = APIRouter()

RESULTS_DIR = "results"
WHO_DATA_PATH = "data/WhoData.xlsx"
os.makedirs(RESULTS_DIR, exist_ok=True)

@router.post("/initialize_audit")
async def initialize_audit(
    full_vin: str = Form(...),
    short_vin: str = Form(...),
    case_spec: str = Form(...),
    variant: str = Form(...),
    engine_number: str = Form(...),
    person_name: str = Form(...),
    person_pno: str = Form(...),
    person_department: str = Form(...),
    components: str = Form(...)  # ✅ Pass JSON string of {interior:[], exterior:[], loose:[]}
):
    try:
        # ✅ 1. Clean up any existing folders with the same VIN
        for folder_name in os.listdir(RESULTS_DIR):
            folder_path_check = os.path.join(RESULTS_DIR, folder_name)
            if os.path.isdir(folder_path_check):
                # Check if folder name starts with the full_vin
                if folder_name.startswith(full_vin) and (' (Ongoing)' in folder_name or ' (Done)' in folder_name):
                    print(f"Deleting existing folder: {folder_name}")
                    shutil.rmtree(folder_path_check)

        # ✅ 2. Create Results Folder
        folder_name = f"{full_vin} (Ongoing)"
        folder_path = os.path.join(RESULTS_DIR, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # ✅ 3. Create InfoBeforeScan.txt
        comps = json.loads(components)
        current_time = datetime.now()
        timestamp = current_time.strftime("%Y-%m-%d %H:%M:%S")
        day_name = current_time.strftime("%A")
        
        lines = [
            f"Audit Started: {timestamp}",
            f"Day: {day_name}",
            "",
            f"Person Pno: {person_pno}",
            f"Person Name: {person_name}",
            f"Person Department: {person_department}",
            f"Full VIN: {full_vin}",
            f"Short VIN: {short_vin}",
            f"Case Spec: {case_spec}",
            f"Variant: {variant}",
            f"Engine Number: {engine_number}",
            "",
            "Components to Check:",
        ]
        for section, items in comps.items():
            lines.append(f"[{section.upper()}]")
            lines.extend([f"- {name}" for name in items])
            lines.append("")

        info_path = os.path.join(folder_path, "InfoBeforeScan.txt")
        with open(info_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        # ✅ 4. Update WhoData.xlsx
        wb = load_workbook(WHO_DATA_PATH)
        sheet = wb.active
        
        # Check if full_vin already exists and update it, otherwise append
        row_found = False
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == full_vin:
                # Update existing row
                row[0].value = full_vin
                row[1].value = short_vin
                row[2].value = person_pno
                row[3].value = person_name
                row[4].value = "Ongoing"
                row_found = True
                break
        
        # If not found, append new row
        if not row_found:
            sheet.append([full_vin, short_vin, person_pno, person_name, "Ongoing"])
        
        wb.save(WHO_DATA_PATH)

        return JSONResponse({
            "status": "success",
            "folder_created": folder_path,
            "info_file": info_path
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
